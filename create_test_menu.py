#!/usr/bin/env python3
import openpyxl

# Create a new workbook
wb = openpyxl.Workbook()
ws = wb.active

# Headers
ws['A1'] = 'Name'
ws['B1'] = 'Ingredients'
ws['C1'] = 'Price'
ws['D1'] = 'Category'

# Test data
test_items = [
    ['Pizza Margherita', 'Tomato, Mozzarella, Basil', 12.50, 'Food'],
    ['Caesar Salad', 'Lettuce, Parmesan, Croutons, Caesar Dressing', 8.90, 'Food'],
    ['Coca Cola', 'Carbonated soft drink', 2.50, 'Drinks'],
    ['Coffee', 'Espresso coffee', 3.00, 'Drinks'],
    ['Chocolate Cake', 'Rich chocolate cake with cream', 6.50, 'Desserts']
]

for i, item in enumerate(test_items, 2):
    ws[f'A{i}'] = item[0]
    ws[f'B{i}'] = item[1]
    ws[f'C{i}'] = item[2]
    ws[f'D{i}'] = item[3]

# Save the file
wb.save('/Users/mos/Desktop/TableLink/test_menu.xlsx')
print("Created test_menu.xlsx")