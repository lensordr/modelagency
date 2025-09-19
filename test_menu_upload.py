#!/usr/bin/env python3
"""
Test script for menu upload functionality
"""

import sys
import os
sys.path.append(os.path.join(os.path.dirname(__file__), 'Restaurant'))

from Restaurant.models import get_db, MenuItem
from Restaurant.setup import process_excel_content
import openpyxl
import io

def create_test_excel():
    """Create a test Excel file with sample menu items"""
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Headers
    ws['A1'] = 'Name'
    ws['B1'] = 'Ingredients'
    ws['C1'] = 'Price'
    ws['D1'] = 'Category'
    
    # Test data
    test_items = [
        ['Pizza Margherita', 'Tomato, Mozzarella, Basil', 12.50, 'Food'],
        ['Caesar Salad', 'Lettuce, Parmesan, Croutons, Caesar Dressing', 8.90, 'Food'],
        ['Coca Cola', 'Carbonated soft drink', 2.50, 'Drinks'],
        ['Coffee', 'Espresso coffee', 3.00, 'Drinks'],
        ['Chocolate Cake', 'Rich chocolate cake with cream', 6.50, 'Desserts']
    ]
    
    for i, item in enumerate(test_items, 2):
        ws[f'A{i}'] = item[0]
        ws[f'B{i}'] = item[1]
        ws[f'C{i}'] = item[2]
        ws[f'D{i}'] = item[3]
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def test_menu_upload():
    """Test the menu upload functionality"""
    print("Testing menu upload functionality...")
    
    # Create test Excel file
    excel_content = create_test_excel()
    print(f"Created test Excel file ({len(excel_content)} bytes)")
    
    # Get database session
    db = next(get_db())
    
    try:
        # Test restaurant ID (use 3 for Mos restaurant)
        restaurant_id = 3
        
        # Check current menu items
        current_items = db.query(MenuItem).filter(MenuItem.restaurant_id == restaurant_id).all()
        print(f"Current menu items for restaurant {restaurant_id}: {len(current_items)}")
        for item in current_items:
            print(f"  - {item.name} ({item.category}) - €{item.price}")
        
        # Process the Excel file
        print("\nProcessing Excel file...")
        process_excel_content(db, excel_content, restaurant_id)
        
        # Check new menu items
        new_items = db.query(MenuItem).filter(MenuItem.restaurant_id == restaurant_id).all()
        print(f"\nNew menu items for restaurant {restaurant_id}: {len(new_items)}")
        for item in new_items:
            print(f"  - {item.name} ({item.category}) - €{item.price}")
        
        # Group by category
        categories = {}
        for item in new_items:
            if item.category not in categories:
                categories[item.category] = []
            categories[item.category].append(item)
        
        print(f"\nCategories found: {list(categories.keys())}")
        for category, items in categories.items():
            print(f"  {category}: {len(items)} items")
        
        print("\n✅ Menu upload test completed successfully!")
        return True
        
    except Exception as e:
        print(f"\n❌ Menu upload test failed: {e}")
        import traceback
        traceback.print_exc()
        return False
    finally:
        db.close()

if __name__ == "__main__":
    test_menu_upload()